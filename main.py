#!/usr/bin/env python
# coding: utf-8

# In[ ]:



import openpyxl
# import time
import xlrd
import re
import math
import xlrd
from decimal import *
import datetime
import binascii,time
import socket

def  readexcel(unu):  #读取execel 中的内容
    workbook=xlrd.open_workbook(r'./mock-data.xls')
#     print(workbook.sheet_names()) #查看excel几个标签页
    sheet1=workbook.sheet_by_name('GPS')
    nrows=sheet1.nrows
#     print(nrows) #查看excel GPS页面几行
    ncols=sheet1.ncols
#     print(ncols)
    lon_lat=sheet1.cell(unu,1).value  #经纬度
    gps_status=sheet1.cell(unu,2).value#上传状态
    run_status=sheet1.cell(unu,0).value#运行状态
    speed=int(sheet1.cell(unu,3).value *10) #运行速度
    mile=int(sheet1.cell(unu,4).value)#运行里程
    flag=sheet1.cell(unu,6).value #上传的是GPS还是到离站
    routon=int(sheet1.cell(unu,5).value) #线路信息
    adtype=sheet1.cell(unu,7).value  #到站离站状态
    adstation=int(sheet1.cell(unu,8).value)#站点编码
    adstno=int(sheet1.cell(unu,9).value) #站点序号
    aduptype=sheet1.cell(unu,10).value #到离站上传方式
    
    return   [nrows,lon_lat,speed,mile,flag,routon,adtype,adstation,adstno,aduptype,run_status]
#*************************************************

#*******************获取系统当前时间  并将时间转化为BCD码
def gettime():
    
    str2=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     print(str2)
    timelist=list(str2)
    for i in timelist:  #处理时间内容

        if i =="-" or i==" " or i==":":
            timelist.remove(i)
        else:
            continue

    dtime=timelist[2:]
    cdtime=map(str,dtime)
    ytime="".join(cdtime)
    return ytime
#***************************************************************

#*************************获取过去时间**************************
def duration():
    
    """
    过去时间 当前时间往前三十秒
    """
    durtime=(datetime.datetime.now()+datetime.timedelta(seconds=-30)).strftime("%Y-%m-%d %H:%M:%S")
    durlist=list(durtime)
    for d  in durlist:
        if d =="-" or d==" " or d==":":
            durlist.remove(d)
        else:
            continue
    durdtime=durlist[2:]
    durcdtime=map(str,durdtime)
    durytime="".join(durcdtime)
    return durytime

#**************************************************************************************************


#****************************格式化 字符串
def mat_def(strcon):
    
    # text='7E0B0200260183399699000051000C8FA00101000F702B0C0101CC9FF2072A0D5A000000000000211011101448000101000000667E'
    text_list=re.findall(".{2}",strcon)
    new_text=" ".join(text_list)
    return new_text
#*********************************异或 BBC计算校验码

def get_bcc(inputStr: str) -> str:
    bcc = 0
    for i in inputStr.split(' '):
        bcc = bcc ^ int(i, 16)
    return f'{bcc:x}'
#**************************************************************
#将字符转化为ascii 十六进制
def to_ascii(cstr):
    mm=''
    for i in cstr:
        k=hex(ord(i))[2:]
       
        mm=''.join([mm,str(k)])
   
    return mm



#*********************************************判断是到站还是离站
def  judge_ad(adnum):
    """
    判断是离站还是到站，离站返回'10',到站返回'01',其他返回01
    """
    adtype3=''
    adtype=readexcel(adnum)[6]
    adtype1=adtype.split('(')
    adtype2=adtype1[0]
    if adtype2=='离站':
        adtype3='02'
        return adtype3
    
        
    else:
        adtype3='01'
        return adtype3
        


#*************************************** 组合报文************************
#**********************************************************************
def combination(nun):
    """
    flagmini=0 组GPS报文
    flagmini=1组到离站报文
    flagmini=2 组的违规包
    flagmini=3 组的是持续违规包
    flagmini=4 组的是考勤包
    """
    final_mes=""
    message_header="7e "  #包头
    message_tail=" 7e" #包尾
    check_code=""
    flagmini=int(readexcel(nun)[4])  #//查看当前行是报站还是GPS
    gps_min=gpsdata(nun)  #GPS
    
    ad_min=addata(nun) #到离站
    
    vio_min=violation(nun) #普通违规
    durvio_min=durviolation(nun) #持续违规
    att_min=attendance(nun) #考勤
    
    if flagmini==0:
#         check_code=
        check_code="".join([" ",get_bcc(gps_min)])
        final_mes="".join([message_header,gps_min,check_code,message_tail])
        print('组合的是GPS包')
        print(final_mes)
    elif flagmini==1:
        
        check_code="".join([" ",get_bcc(ad_min)])
        print(check_code)
        final_mes="".join([message_header,ad_min,check_code,message_tail])
        print("组合的是到离站包")
#         print(ad_min)
        print(final_mes)
    elif flagmini==2:
        check_code="".join([" ",get_bcc(vio_min)])
        final_mes="".join([message_header,vio_min,check_code,message_tail])
        print('组合的是超速违规包')
        print(final_mes)
    
    
    
    elif flagmini==3:
        
        check_code="".join([" ",get_bcc(durvio_min)])
        final_mes="".join([message_header,durvio_min,check_code,message_tail])
        print('组合的是超速持续违规包')
        print(final_mes)
    elif flagmini==4:
        
        check_code="".join([" ",get_bcc(att_min)])
        final_mes="".join([message_header,att_min,check_code,message_tail])
        print('组合的是考勤包')
        print(final_mes)
        
        
    else:
        print("啥都不是")
    return  final_mes
        
#************************************** 将报文转化为16进制内容用于发送***********
def  test_hex(test_bbbb:str)->str:
    d=b''
    for i in test_bbbb.split(' '):
        ia=int(i,16)
#         print(ia)
        mm=ia.to_bytes(1,byteorder='little', signed=False)
        d=b"".join([d,mm])
#         print(d)
    print(d)
    return d
#*************************************************************************************
#高德坐标转化为84坐标系   GCJ02  ->WGS84
x_pi=3.14159265358979324 * 3000.0 / 180.0
pi = 3.1415926535897932384626  # π
a = 6378245.0  # 长半轴
ee = 0.00669342162296594323  # 扁率

def handlelng_lat(unu):
    strlnglat=readexcel(unu)[1]
    ys=re.findall(r'[(](.*?)[)]',strlnglat)
    yss=ys[0].split(',')
    lat1 =float (yss[0])#纬度
    lng1=float (yss[1]) #经度
#     print(lat1,lng1)
    return[lat1,lng1]
def gcj02towgs84(lng, lat):
    """
    GCJ02(火星坐标系)转GPS84
    :param lng:火星坐标系的经度
    :param lat:火星坐标系纬度
    :return:
    """
    dlat = transformlat(lng - 105.0, lat - 35.0)
    dlng = transformlng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)
    mglat = lat + dlat
    mglng = lng + dlng
    strlng=format(lng * 2 - mglng,'.6f')
    strlat=format(lat * 2 - mglat,'.6f')
    #去掉小数点并已整数输出
    relng=int((float(strlng))*1000000)
    relat=int((float(strlat))*1000000)
      
        
    return [relng, relat]



def transformlat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat +         0.1 * lng * lat + 0.2 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * pi) + 40.0 *
            math.sin(lat / 3.0 * pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * pi) + 320 *
            math.sin(lat * pi / 30.0)) * 2.0 / 3.0
    return ret

def transformlng(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng +         0.1 * lng * lat + 0.1 * math.sqrt(math.fabs(lng))
    ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 *
            math.sin(2.0 * lng * pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * pi) + 40.0 *
            math.sin(lng / 3.0 * pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * pi) + 300.0 *
            math.sin(lng / 30.0 * pi)) * 2.0 / 3.0
    return ret
# gcj02towgs84(handlelng_lat(2)[1],handlelng_lat(2)[0])
#********************************************************************
def  trantohex(abc,bn):    #将数值转化为16 进制，根据位数生成对应的内容
    abc2=hex(abc)
    abc3=list(abc2)
    abc4=abc3[2:]  #去掉0x
#     print(abc4)
    while len(abc4)<bn:  #判断 转化后的报文长度，是否小于bn位，小于bn位自动在前面补0
        abc4.insert(0,'0')
#     print(abc4)
    abc5=map(str,abc4)
    reabc="".join(abc5)
    return reabc  
#**********************************************************
# lngtt=gcj02towgs84(handlelng_lat(2)[1],handlelng_lat(2)[0])[0]  #获得当前行的经度
# # newlngtt=lngtt.replace('.','')
# # print(newlngtt)
# print(trantohex(lngtt,8))
# lattt=gcj02towgs84(handlelng_lat(2)[1],handlelng_lat(2)[0])[1] #获得当前行的纬度
# trantohex(lattt,8)

#*********************************GPS运行状态************************
def gpstype_def(tnun):
    """
    tnum  第几行
    GPS状态
    上行	0x01
	下行	0x02
	环行	0x03
	停主站	0x04
	停副站	0x05
	保留	0x06~0x0F
	自定义	0x20~0x7F
	出场	0x80
	进场	0x81
	加油	0x82
	加气	0x83
	充电	0x84
	小修	0x85
	大修	0x86
	一保	0x87
	二保	0x88
	三保	0x89
	放空	0x8A
	停场	0x8B
	保留	0x8C~0x9F
	自定义	0xA0~0xFF

    
    
    """
    tnum=readexcel(tnun)[10]
    print('是啥呀',tnum)
    typevalue=''
    
    if tnum=='上行':
        typevalue='01'
    elif tnum=='下行':
        typevalue='02'
    elif tnum=='环行':
        typevalue='03'
    elif tnum=='停主站':
        typevalue='04'
    elif tnum=='停副站':
        typevalue='05'
    elif tnum=='出场':
        typevalue='80'
    elif tnum=='进场':
        typevalue='81'
    elif tnum=='加油':
        typevalue='82'
    elif tnum=='加气':
        typevalue='83'
    elif tnum=='充电':
        typevalue='84'
    elif tnum=='小修':
        typevalue='85'
    elif tnum=='大修':
        typevalue='86'
    elif tnum=='一保':
        typevalue='87'
    elif tnum=='二保':
        typevalue='88'
    elif tnum=='三保':
        typevalue='89'
    elif tnum=='放空':
        typevalue='8A'
    elif tnum=='停场':
        typevalue='8B'
    
    
    else :
        typevalue='01'
    return typevalue

#********************************GPS报文封装*************************
def gpsdata(nun):
    """

    nun  第几行,

    
    
    
    
    """
    
    

    GPSstr=''
    #******** 包头*******************
    #消息 id 第二个位置
    msgid='0200'
  

    #消息长度（消息对象属性） GPS 定长67 转移后0043 第三个位置
    msglg='0043'


    #终端手机号 固定018339969900 第四个位置
    tvno='018339969900'


    #消息流水号 ,可以根据  第五个位置  ，长度4  取excel所在的行数
    msgno=trantohex(45820,4)
   
    #*************包体**********************
    #报警状态  默认000000000 第6个位置，长度8
    armno='00000000'

    #状态 默认00000000 ，第7位，长度8        
    armid='00000002'
  
    #纬度度  07298109  ,第8位，长度8 取excel经纬度中的经度转义后的内容
    
    gpslngtt=gcj02towgs84(handlelng_lat(nun)[1],handlelng_lat(nun)[0])[1] #获得当前行的纬度
    gpslng=trantohex(gpslngtt,8)
   
    #经度  01cd9d56   ,第9位，长度8取excel经纬度中的纬度转义后的内容
    gpslattt=gcj02towgs84(handlelng_lat(nun)[1],handlelng_lat(nun)[0])[0] #获得当前行的经度
    gpslat=trantohex(gpslattt,8)
    

    #高度 默认0000  第10项，长度4为
    gpshight='0000'
    
    #速度 读excel 速度 *10   第11项 长度4
    gpssseed=readexcel(nun)[2]
        

    gpsspeed =trantohex(gpssseed,4)
    

    #方向 默认0000， 第12项 长度4位
    gpsdir='0000'
   
    #时间  获取系统当前时间， 第13项，长度12
    gpstime=gettime()
    
    #附加信息  标识01 附加长度04    第14项   0104  
    gpsadd01='0104'
    
    #附加信息 里程 长度8位    第15项 里程取excel 里程列 
    gpssmile=readexcel(nun)[3]
    gpsmile=trantohex(gpssmile,8)
   

    #附加 标识 02 附加长度02   第16项   0202
    gpsadd02='0202'
    

    #附加信息 油量  长度4    第 17 项 使用默认值0000
    gpsoil='0000'
   
    #附加信息标志03  附加长度02 ，0302 第18项
    gpsadd03='0302'
   

    #附加消息标志 速度  第19项， 长度4  默认0
    
    gpsspeed2=trantohex(230,4)
    

    #附加信息 标志04 附加长度02  0402  第20项
    gpsadd04='0402'
   
    #附加消息 报警事件id  长度4 第21项
    gpsalarm=trantohex(0,4)
    


    #附加信息 标志14 长度标志04 1404   第22 项
    gpsadd14='1404'
   

    #附加消息 视频报警事件id  长度8 第23项
    gpsvideoalarm=trantohex(0,8)
   


    #附加信息 标志15 长度标志04 1504   第24项
    gpsadd15='1504'
    

    #附加信息 驾驶行为分析  第25项
    gpsemp=trantohex(0,8)
    

    #附加信息 标志16 长度标志04 1604   第26项
    gpsadd16='1604'
    

    #附加信息 线路编码  第27项
    gpsrounos=readexcel(nun)[5]
    gpsrouno=trantohex(gpsrounos,8)
    

    #附加信息 标志17 长度标志01 1701   第28项
    gpsadd17='1701'
    

    #附加信息 业务类型  第29项 长度2取excel经纬度中的运行方向转义后的内容？
    gpstype=gpstype_def(nun)
#     gpstype=trantohex(gpstype2,2)
    

    GPSstr="".join([msgid,msglg,tvno,msgno,armno,armid,gpslng,gpslat,gpshight,gpsspeed,gpsdir,gpstime,gpsadd01,gpsmile,gpsadd02,gpsoil,gpsadd03,gpsspeed2,gpsadd04,gpsalarm,gpsadd14,gpsvideoalarm,gpsadd15,gpsemp,gpsadd16,gpsrouno,gpsadd17,gpstype])

    
    return mat_def(GPSstr)



# print(gpsdata(1))
# get_bcc(gpsdata(1))
# if __name__ == "__main__":
    
#     extranslate=combination(8)  #将excel 中的内容转化为有效报文
#     aftranslate=test_hex(extranslate)  #将报文转化为16进制可以发送的报文
#     print(aftranslate)
    





#到离站区域
#***************************************************************************

def addata(num):
    """
    到离站报文
    anun,
    anun  第几行,

    arlng  经度,
    arlat  纬度,
    arroutno 线路编码,
    arspeed 速度,
    ar
    extype 状态
    
    """
    Ardrstr=""
    #******************************************包头*******************
    #消息id 在第二位置
    admgid='0B02'
    
    #消息长度（消息对象属性） 第三个位置  考虑可变？
    admsglg='0026'


    #终端手机号 固定018339969900 第四个位置
    adtvno='018339969900'


    #消息流水号 ,可以根据  第五个位置  ，长度4  取excel所在的行数
#     admsgno=trantohex(num,4)
    admsgno='0020'

    #**************************************包体***************************
    
    #********************************线路编码****************************
    adrounos=readexcel(num)[5]
    print(adrounos)
    adrouno=trantohex(adrounos,8)
    
    
    #*******************************到离站类型*读取excel中的 内容* 是到站还是离站 到站01，离站10？ ****************************
    adtypef=judge_ad(num)
    
    #*****************************业务类型（上行，下行)************************
    addictypef=gpstype_def(num)
    #******************************场站站点编号* 读取excel中 站点编号* 长度8位***************************
    stationf1=readexcel(num)[7]
    stationf=trantohex(stationf1,8)
    #*******************************车站序号*  读取excel中 站点序号，长度2位 ********************************
    adstonf1=readexcel(num)[8]
    adstonf=trantohex(adstonf1,2)
    #**********************标志字段 自动正常还是补发  * 读取excel 中的上报方式 长度2位 在拓展，先默认自动 01 ********************************
    aduptype='00'
    
    #*************************纬度信息***********************************（可能混了，就先这样吧）************
    adlngtt=gcj02towgs84(handlelng_lat(num)[1],handlelng_lat(num)[0])[1] #获得当前行的纬度
    adlng=trantohex(adlngtt,8)
    
    #**************************经度信息**********************************（可能混了，就先这样吧）***********
    adlattt=gcj02towgs84(handlelng_lat(num)[1],handlelng_lat(num)[0])[0] #获得当前行的经度
    adlat=trantohex(adlattt,8)
    
    #****************************高程******默认0 长度4********************************
    adhight='0000'
    
    #*****************************车速******默认0 ，长度4******************************
    adspeed='0000'
    #************************************方向角 默认0 长度4*************
    addicag='0000'
    
    #*****************************系统时间** 长度12****************************
    adtime=gettime()
    #**********************************当前乘客数** 长度4*********************************
    adpassenger='0001'
    #车门数 01
    addnoniu='01'
    #车门开关情况************000000
    addnoniunu='000000'
    
    #************************组合到离站包
    Ardrstr="".join([admgid,admsglg,adtvno,admsgno,adrouno,adtypef,addictypef,stationf,adstonf,aduptype,adlng,adlat,adhight,adspeed,addicag,adtime,adpassenger,addnoniu,addnoniunu])
    
    return  mat_def(Ardrstr)
#******************************************************************


#******************************违规**************************************************


#***************************************************************************

def violation(num):
    """
    违规上报
    vnun,
    vnun  第几行,

    vrlng  经度,
    vrlat  纬度,
    vrroutno 线路编码,
    vrspeed 速度,
    vr
    extype 状态
    
    """
    viostr=""
    #******************************************包头*******************
    #消息id 在第二位置
    vioid='0B04'
    
    #消息长度（消息对象属性） 第三个位置  考虑可变？
    violg='001E'


    #终端手机号 固定018339969900 第四个位置
    viotvno='018339969900'


    #消息流水号 ,可以根据  第五个位置  ，长度4  取excel所在的行数
#     admsgno=trantohex(num,4)
    viomsgno='0020'

    #**************************************包体***************************
    
    #********************************线路编码****************************
    viorounos=readexcel(num)[5]
    print(viorounos)
    viorouno=trantohex(viorounos,8)
    
    
    #*******************************违规类型 默认 01  超速 ****************************
    viotype='01'
    
    #*****************************违规值 默认55************************
    viovalue='157c'
    #***************************违规标准******************************
    viostandard='1194'
    #*******************************车站序号*  读取excel中 站点序号，长度2位 ********************************
    viostonf1=readexcel(num)[8]
    viostonf=trantohex(viostonf1,2)
    #*************************纬度信息***********************************（可能混了，就先这样吧）************
    violngtt=gcj02towgs84(handlelng_lat(num)[1],handlelng_lat(num)[0])[1] #获得当前行的纬度
    violng=trantohex(violngtt,8)
    
    #**************************经度信息**********************************（可能混了，就先这样吧）***********
    violattt=gcj02towgs84(handlelng_lat(num)[1],handlelng_lat(num)[0])[0] #获得当前行的经度
    violat=trantohex(violattt,8)
    
    #****************************高程******默认0 长度4********************************
    viohight='0000'
    
    #*****************************车速******默认0 ，长度4******************************
    viospeed='0000'
    #************************************方向角 默认0 长度4*************
    viodicag='0000'
    
    #*****************************系统时间** 长度12****************************
    viotime=gettime()
    #**********************标志字段 自动正常还是补发  * 读取excel 中的上报方式 长度2位 在拓展，先默认自动 01 ********************************
    viouptype='00'

    
    
    #************************组合违规包
    viostr="".join([vioid,violg,viotvno,viomsgno,viorouno,viotype,viovalue,viostandard,violng,violat,viohight,viospeed,viodicag,viotime,viouptype])
    
    return  mat_def(viostr)
#*****************************持续违规*************************************

def durviolation(num):
    """
    持续违规上报
    vnun,
    vnun  第几行,

    vrlng  经度,
    vrlat  纬度,
    vrroutno 线路编码,
    vrspeed 速度,
    vr
    extype 状态
    
    """
    durviostr=""
    #******************************************包头*******************
    #消息id 在第二位置
    durvioid='0B04'
    
    #消息长度（消息对象属性） 第三个位置  考虑可变？
    durviolg='0040'


    #终端手机号 固定018339969900 第四个位置
    durviotvno='018339969900'


    #消息流水号 ,可以根据  第五个位置  ，长度4  取excel所在的行数
#     admsgno=trantohex(num,4)
    durviomsgno='0020'

    #**************************************包体***************************
    
    #********************************线路编码****************************
    durviorounos=readexcel(num)[5]
    print(durviorounos)
    durviorouno=trantohex(durviorounos,8)
    
    
    #*******************************违规类型 默认 01  超速 ****************************
    durviotype='01'
    
    #*****************************违规值 默认55************************
    durviovalue='157c'
    #***************************违规标准******************************
    durviostandard='1194'
    
    #*************************纬度信息***********************************（可能混了，就先这样吧）************
    durviolngtt=gcj02towgs84(handlelng_lat(num)[1],handlelng_lat(num)[0])[1] #获得当前行的纬度
    durviolng=trantohex(durviolngtt,8)
    
    #**************************经度信息**********************************（可能混了，就先这样吧）***********
    durviolattt=gcj02towgs84(handlelng_lat(num)[1],handlelng_lat(num)[0])[0] #获得当前行的经度
    durviolat=trantohex(durviolattt,8)
    
    #****************************高程******默认0 长度4********************************
    durviohight='0000'
    
    #*****************************车速******默认0 ，长度4******************************
    durviospeed='0000'
    #************************************方向角 默认0 长度4*************
    durviodicag='0000'
    
    #*****************************系统时间** 长度12****************************
    durviotime=gettime()
    #**********************标志字段 自动正常还是补发  * 读取excel 中的上报方式 长度2位 在拓展，先默认自动 01  后面多补两个0 用于加到附加信息上 ********************************
    durviouptype='00'
    
    
    #******持续违规***************************************************************
    additional=''
    
    isdurvio='01' #是否持续超速
    durviostarttime=duration() #超速开始时间
    durvioendtime=gettime()#超速结束时间
    durvioavgspeed='157c'#平均速度
    durviomaxspeed='19c8'#最高速度
    durviobelat=durviolng  #开始纬度
    durviobelon=durviolat  #开始经度
    durvioaflat=durviolng  #开始纬度
    durvioaflon=durviolat  #开始经度
    
    
    
    additional="".join([isdurvio,durviostarttime,durvioendtime,durvioavgspeed,durviomaxspeed,durviobelat,durviobelon,durvioaflat,durvioaflon])
    readditional=additional
#     print(readditional)
    #附加内容结束
    strend='00' #字符串结束
    
    #************************组合持续违规包
    durviostr="".join([durvioid,durviolg,durviotvno,durviomsgno,durviorouno,durviotype,durviovalue,durviostandard,durviolng,durviolat,durviohight,durviospeed,durviodicag,durviotime,durviouptype,readditional,strend])
#     print(durviostr)
    return  mat_def(durviostr)
#******************************************************************


#*********************考勤******************************************************

def attendance(num):
    """
	7E 
	0b05  消息类型
	0015 消息长度
	018506332469 手机编号
	01ba  消息流水
	00001f61 线路编号
	3337616303561640
	021210140052  考勤时间
	04   考勤类型
	01  考勤方式


	64 校验码
    
    """
    attendstr=""
    #******************************************包头*******************
    #消息id 在第二位置
    attendid='0B05'
    
    #消息长度（消息对象属性） 第三个位置  考虑可变？
    attendlg='0015'


    #终端手机号 固定018339969900 第四个位置
    attendtvno='018339969900'


    #消息流水号 ,可以根据  第五个位置  ，长度4  取excel所在的行数

    attendmsgno='0020'

    #**************************************包体***************************
    
    #********************************线路编码****************************
    attendrounos=readexcel(num)[5]
    print(attendrounos)
    attendrouno=trantohex(attendrounos,8)
    #********************************员工编号
    attempon='363735353166616400'
   
    
    #*****************************考勤时间** 长度12****************************
    attendtime=gettime()
    #**********************考勤类型 *上班01，下班02，签到03 签退 04*******************************
    atttype='04'
    #**********************************考勤方式    员工卡 01，工号 02
    attmode='02'

    
    
    #************************组合违规包
    attendstr="".join([attendid,attendlg,attendtvno,attendmsgno,attendrouno,attempon,attendtime,atttype,attmode])
    
    return  mat_def(attendstr)

























# # #发送测试**********************************************7776163**********************
urows=123
print(readexcel(1)[0]) #excel 行数
host ='101.37.135.193'
port = 16718
client = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
client.settimeout(5)
client.connect((host,port))
register_test="7E 01 00 00 1a 01 83 39 96 99 00 00 06 00 1f 00 48 43 43 43 42 42 61 6e 64 72 6f 69 64 30 00 00 00 00 76 a7 a3 00 00 84 7E" #注册
authentication_test="7E 01 02 00 0C 01 83 39 96 99 00 00 06 35 34 31 32 39 31 31 36 31 00 00 00 81 7E" #鉴权
heartbeat_test="7E 00 02 00 00 01 83 39 96 99 00 00 2D 9B 7E"
heartb=test_hex(heartbeat_test)
reg=test_hex(register_test)

aut=test_hex(authentication_test)

client.send(reg) #发送注册
time.sleep(5)
client.send(aut)#发送鉴权


# readexcel(1)[0]
#每5秒钟发送一次 GPS
while (urows<readexcel(1)[0]):
    client.send(heartb) #发送心跳
    time.sleep(3)
    print("第",urows)
    print(readexcel(urows)[1])
    extranslate=combination(urows)  #将excel 中的内容转化为有效报文
    aftranslate=test_hex(extranslate)  #将报文转化为16进制可以发送的报文
    client.send(aftranslate)    #发送GPS
    urows=urows+1


# In[5]:


#
def to_ascii(cstr):
    mm=''
    for i in cstr:
        k=hex(ord(i))[2:]
       
        mm=''.join([mm,str(k)])
   
    return mm
    
        
    

teststr='398903231'
print(to_ascii(teststr))


# In[4]:



#将16进制字符转化为 10进制字符
import codecs

codecs.decode('3337616303561640', "hex").decode('utf-8')


# In[8]:


aaaaaa='00033086010000119401CD9F21072915D70000000000002112160953380030313231313231353134343030373231313231353134343030373031464530323034'


print(len(aaaaaa))


# In[20]:


cccc='30313231313231363131303333323231313231363131303430323135376331396338'
print(len(cccc))


# In[13]:


import datetime
def gettime():
    
    str2=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(str2)
    timelist=list(str2)
    for i in timelist:  #处理时间内容

        if i =="-" or i==" " or i==":":
            timelist.remove(i)
        else:
            continue

    dtime=timelist[2:]
    cdtime=map(str,dtime)
    ytime="".join(cdtime)
    return ytime
gettime()


# In[16]:


import datetime
print ((datetime.datetime.now()+datetime.timedelta(seconds=-30)).strftime("%Y-%m-%d %H:%M:%S"))


# In[17]:


import datetime
def duration():
    
    """
    过去时间
    """
    durtime=(datetime.datetime.now()+datetime.timedelta(seconds=-30)).strftime("%Y-%m-%d %H:%M:%S")
    durlist=list(durtime)
    for d  in durlist:
        if d =="-" or d==" " or d==":":
            durlist.remove(d)
        else:
            continue
    durdtime=durlist[2:]
    durcdtime=map(str,durdtime)
    durytime="".join(durcdtime)
    return durytime
duration()       


# In[ ]:




